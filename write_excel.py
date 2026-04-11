#!/usr/bin/env python3
"""Writes job application data (JSON) to a formatted Excel spreadsheet with Job ID tracking."""

import json
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATUS_COLORS = {
    "applied": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "interview": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "offer": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "accepted": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "follow-up": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "no response": PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),
    "other": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

HEADERS = ["Job ID", "Company", "Role", "Status", "Source", "Date Applied", "Last Updated", "Email Subject", "Status History", "Notes"]
COL_WIDTHS = {"A": 12, "B": 25, "C": 30, "D": 16, "E": 16, "F": 14, "G": 14, "H": 40, "I": 35, "J": 40}


def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="medium", color="1F3864"))

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 25
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def get_status_fill(status):
    status_lower = status.lower()
    for key, fill in STATUS_COLORS.items():
        if key in status_lower:
            return fill
    return STATUS_COLORS["other"]


def format_status_history(history):
    if not history or not isinstance(history, list):
        return ""
    return " -> ".join(f"{h.get('status', '?')} ({h.get('date', '?')})" for h in history)


def write_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    style_header(ws)

    data_font = Font(size=11, name="Calibri")
    id_font = Font(size=11, name="Calibri", bold=True, color="2F5496")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    # Sort by last_updated descending, then date descending
    try:
        data.sort(key=lambda x: x.get("last_updated", x.get("date", "")), reverse=True)
    except Exception:
        pass

    for row_idx, entry in enumerate(data, 2):
        # Job ID
        id_cell = ws.cell(row=row_idx, column=1, value=entry.get("job_id", ""))
        id_cell.font = id_font
        id_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Company
        ws.cell(row=row_idx, column=2, value=entry.get("company", "Unknown")).font = data_font

        # Role (hyperlinked to job posting if URL available)
        role = entry.get("role", "Unknown")
        job_url = entry.get("job_url", "")
        role_cell = ws.cell(row=row_idx, column=3, value=role)
        if job_url:
            role_cell.hyperlink = job_url
            role_cell.font = Font(size=11, name="Calibri", color="0563C1", underline="single")
        else:
            role_cell.font = data_font

        # Status (color-coded)
        status = entry.get("status", "Other")
        status_cell = ws.cell(row=row_idx, column=4, value=status)
        status_cell.font = Font(size=11, name="Calibri", bold=True)
        status_cell.fill = get_status_fill(status)
        status_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Source
        ws.cell(row=row_idx, column=5, value=entry.get("source", "")).font = data_font

        # Date Applied
        ws.cell(row=row_idx, column=6, value=entry.get("date", "")).font = data_font

        # Last Updated
        ws.cell(row=row_idx, column=7, value=entry.get("last_updated", "")).font = data_font

        # Email Subject
        ws.cell(row=row_idx, column=8, value=entry.get("subject", "")).font = data_font
        ws.cell(row=row_idx, column=8).alignment = wrap_alignment

        # Status History
        history_str = format_status_history(entry.get("status_history", []))
        ws.cell(row=row_idx, column=9, value=history_str).font = Font(size=10, name="Calibri", color="666666")
        ws.cell(row=row_idx, column=9).alignment = wrap_alignment

        # Notes
        ws.cell(row=row_idx, column=10, value=entry.get("notes", "")).font = data_font
        ws.cell(row=row_idx, column=10).alignment = wrap_alignment

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row + Job ID column
    ws.freeze_panes = "B2"

    # ── Summary sheet ─────────────────────────────────────────
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Job Application Summary").font = Font(bold=True, size=14, name="Calibri")
    summary_ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=11, name="Calibri", color="666666")
    summary_ws.cell(row=3, column=1, value=f"Total Applications Tracked: {len(data)}").font = Font(size=11, name="Calibri")

    # Status breakdown
    status_counts = {}
    for entry in data:
        s = entry.get("status", "Other")
        status_counts[s] = status_counts.get(s, 0) + 1

    summary_ws.cell(row=5, column=1, value="Status").font = Font(bold=True, size=11, name="Calibri")
    summary_ws.cell(row=5, column=2, value="Count").font = Font(bold=True, size=11, name="Calibri")

    for i, (status, count) in enumerate(sorted(status_counts.items()), 6):
        summary_ws.cell(row=i, column=1, value=status).font = Font(size=11, name="Calibri")
        cell = summary_ws.cell(row=i, column=2, value=count)
        cell.font = Font(size=11, name="Calibri")

    total_row = 6 + len(status_counts)
    summary_ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True, size=11, name="Calibri")
    summary_ws.cell(row=total_row, column=2, value=len(data)).font = Font(bold=True, size=11, name="Calibri")

    # Recent activity
    recent_row = total_row + 2
    summary_ws.cell(row=recent_row, column=1, value="Recent Activity").font = Font(bold=True, size=12, name="Calibri")
    for i, entry in enumerate(data[:10], recent_row + 1):
        job_id = entry.get("job_id", "")
        company = entry.get("company", "")
        status = entry.get("status", "")
        last = entry.get("last_updated", "")
        summary_ws.cell(row=i, column=1, value=f"{job_id} - {company}").font = Font(size=11, name="Calibri")
        summary_ws.cell(row=i, column=2, value=status).font = Font(size=11, name="Calibri")
        summary_ws.cell(row=i, column=3, value=last).font = Font(size=11, name="Calibri", color="666666")

    summary_ws.column_dimensions["A"].width = 35
    summary_ws.column_dimensions["B"].width = 15
    summary_ws.column_dimensions["C"].width = 14

    wb.save(output_path)
    print(f"  Saved {len(data)} job applications to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: write_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], "r") as f:
        content = f.read().strip()
        # Strip markdown code fences if present
        if content.startswith("```"):
            content = "\n".join(content.split("\n")[1:])
        if content.endswith("```"):
            content = "\n".join(content.split("\n")[:-1])
        data = json.loads(content)

    write_excel(data, sys.argv[2])
